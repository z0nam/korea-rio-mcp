"""Induced-coefficient extraction and loading (region-agnostic).

Ported from 26p17 ``extract_induce_coefficients.py``. The Jeju-specific
``_jeju`` / ``_outside`` column names are generalized to ``_in_region`` /
``_out_region``; sheet names and the region/national subtotal rows come from
:mod:`rio_mcp.engine.regions`.

Column-sum semantics (unchanged from the source): for each sector column ``j``,
the sum over all producing sectors of the induced effect of 1 unit of final
demand in ``j``. In-region = the region-subtotal row; out-of-region leakage =
national total − in-region.

Units: production / value-added = 원/원; employment = 명/10억원.
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd

from .regions import RegionTableLayout

METRICS = ("production", "value_added", "employment")

# Generic coefficient columns the engine consumes.
COEF_COLUMNS = [
    "multiplier_production_in_region", "multiplier_production_out_region",
    "multiplier_value_added_in_region", "multiplier_value_added_out_region",
    "multiplier_employment_in_region", "multiplier_employment_out_region",
]

# Legacy 26p17 column names -> generic, so old Jeju CSVs load unchanged.
_LEGACY_RENAME = {
    "multiplier_production_jeju": "multiplier_production_in_region",
    "multiplier_production_outside": "multiplier_production_out_region",
    "multiplier_value_added_jeju": "multiplier_value_added_in_region",
    "multiplier_value_added_outside": "multiplier_value_added_out_region",
    "multiplier_employment_jeju": "multiplier_employment_in_region",
    "multiplier_employment_outside": "multiplier_employment_out_region",
}


def _column_sums(ws, layout: RegionTableLayout, metric: str) -> pd.DataFrame:
    codes, names = {}, {}
    for c in range(layout.data_col_start, layout.data_col_end + 1):
        code = ws.cell(row=layout.sector_code_row, column=c).value
        name = ws.cell(row=layout.sector_name_row, column=c).value
        codes[c] = str(code).zfill(2) if code is not None else None
        names[c] = name

    rows = []
    for c in range(layout.data_col_start, layout.data_col_end + 1):
        in_region = ws.cell(row=layout.region_subtotal_row, column=c).value
        in_region = float(in_region) if isinstance(in_region, (int, float)) else 0.0
        national = ws.cell(row=layout.national_total_row, column=c).value
        national = float(national) if isinstance(national, (int, float)) else 0.0
        out_region = max(0.0, national - in_region)
        rows.append({
            "sector_code": codes[c],
            "sector_name": names[c],
            f"multiplier_{metric}_in_region": in_region,
            f"multiplier_{metric}_out_region": out_region,
        })
    return pd.DataFrame(rows)


def extract_coefficients(xlsx_path: str | Path, layout: RegionTableLayout) -> pd.DataFrame:
    """Extract induced coefficients from a BOK regional IO workbook.

    Returns a DataFrame with ``sector_code``, ``sector_name``, the six generic
    multiplier columns, unit/metadata columns, ready to cache as CSV.
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    frames = {}
    for metric in METRICS:
        sheet = layout.sheets[metric]
        frames[metric] = _column_sums(wb[sheet], layout, metric)

    out = frames["production"][["sector_code", "sector_name"]].copy()
    for metric in METRICS:
        for scope in ("in_region", "out_region"):
            col = f"multiplier_{metric}_{scope}"
            out[col] = frames[metric][col]
    out["unit_production"] = "원/원"
    out["unit_value_added"] = "원/원"
    out["unit_employment"] = "명/10억원"
    out["region"] = layout.region
    out["table_year"] = layout.table_year
    out["classification"] = layout.classification
    return out


def load_coefficients(csv_path: str | Path) -> pd.DataFrame:
    """Load a cached coefficient CSV, indexed by ``sector_code``.

    Accepts both the generic column names and the legacy 26p17 ``_jeju`` /
    ``_outside`` names (auto-renamed), so existing Jeju tables work as-is.
    """
    df = pd.read_csv(csv_path, dtype={"sector_code": str})
    df = df.rename(columns=_LEGACY_RENAME)
    df["sector_code"] = df["sector_code"].str.zfill(2)
    missing = [c for c in COEF_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f"Coefficient file {csv_path} missing columns: {missing}")
    return df.set_index("sector_code")[COEF_COLUMNS]
