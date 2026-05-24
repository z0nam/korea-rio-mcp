"""Multiregional coefficient extraction from KOSIS/BOK regional IO workbooks.

The regional IO 유발계수 workbooks (KOSIS table DT_301010_FILE20201, 지역표) lay
out an inter-regional matrix: rows = region × sector, columns = region × sector,
with a final 전국 (all-region) total row. Unlike the Jeju-only cut, there is no
pre-computed per-region subtotal row, so the in-region effect is obtained by
**summing the target region's own row-block** for each of its demand columns:

* in-region[s_j]  = Σ over rows of region R of M[row, (R, s_j)]
* national[s_j]   = M[전국 total row, (R, s_j)]
* out-region[s_j] = national − in-region   (leakage to other regions)

Validated against the 26p17 Jeju 중분류 coefficients (max abs diff ~2e-16).

Each metric (production / value-added / employment) is a separate workbook, so
:func:`extract_region` reads one metric-file at a time; :func:`build_region_table`
merges metrics into the engine's generic coefficient CSV.
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd

# 17 regions as labelled in the workbooks (row/column label column).
REGIONS_17 = [
    "서울", "인천", "경기", "대전", "세종", "충북", "충남", "광주", "전북",
    "전남", "대구", "경북", "부산", "울산", "경남", "강원", "제주",
]
NATIONAL_LABEL = "전국"

# Header layout shared by the regional 유발계수 workbooks.
REGION_LABEL_ROW = 5     # 1-based: row holding region names across columns
SECTOR_CODE_ROW = 6
SECTOR_NAME_ROW = 7
LABEL_COL = 1            # column holding the row-region label (1-based)


def _load_grid(xlsx_path: str | Path, sheet: str | None = None) -> list[list]:
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _index_blocks(grid: list[list]):
    """Return (row_blocks, col_blocks, national_row_idx, sector_meta).

    row_blocks/col_blocks: {region -> [0-based indices]}. sector_meta is keyed
    by 0-based column index -> (code, name) for that column.
    """
    ncol = max(len(r) for r in grid)
    reg_row = grid[REGION_LABEL_ROW - 1]
    code_row = grid[SECTOR_CODE_ROW - 1]
    name_row = grid[SECTOR_NAME_ROW - 1]

    col_blocks: dict[str, list[int]] = {}
    sector_meta: dict[int, tuple[str, str]] = {}
    for c in range(ncol):
        rg = reg_row[c] if c < len(reg_row) else None
        if rg:
            col_blocks.setdefault(rg, []).append(c)
            code = code_row[c] if c < len(code_row) else None
            name = name_row[c] if c < len(name_row) else None
            if code is not None:
                sector_meta[c] = (str(code).zfill(2), name)

    row_blocks: dict[str, list[int]] = {}
    for i, r in enumerate(grid):
        a = r[LABEL_COL - 1] if len(r) >= LABEL_COL else None
        if a:
            row_blocks.setdefault(a, []).append(i)

    nat_idx = row_blocks[NATIONAL_LABEL][0]
    return row_blocks, col_blocks, nat_idx, sector_meta


def extract_region(xlsx_path: str | Path, region: str, metric: str) -> pd.DataFrame:
    """Extract one region's in/out-region induced coefficients for one metric.

    Returns columns: sector_code, sector_name,
    multiplier_{metric}_in_region, multiplier_{metric}_out_region.
    """
    grid = _load_grid(xlsx_path)
    row_blocks, col_blocks, nat_idx, sector_meta = _index_blocks(grid)
    if region not in row_blocks or region not in col_blocks:
        raise KeyError(f"Region '{region}' not found in {xlsx_path}")

    region_rows = row_blocks[region]
    rec = []
    for c in col_blocks[region]:
        code, name = sector_meta[c]
        in_region = sum(grid[i][c] for i in region_rows
                        if c < len(grid[i]) and isinstance(grid[i][c], (int, float)))
        nat = grid[nat_idx][c] if c < len(grid[nat_idx]) and isinstance(grid[nat_idx][c], (int, float)) else 0.0
        rec.append({
            "sector_code": code,
            "sector_name": name,
            f"multiplier_{metric}_in_region": in_region,
            f"multiplier_{metric}_out_region": max(0.0, nat - in_region),
        })
    return pd.DataFrame(rec).sort_values("sector_code").reset_index(drop=True)


def build_region_table(region: str, metric_files: dict[str, str | Path],
                       table_year: int, classification: str) -> pd.DataFrame:
    """Merge per-metric extractions for one region into a coefficient table.

    ``metric_files``: {"production": path, "value_added": path,
    "employment": path}. Missing metrics are filled with NaN columns so the
    table shape is stable (e.g. employment absent from the KOSIS regional set).
    """
    base = None
    for metric in ("production", "value_added", "employment"):
        path = metric_files.get(metric)
        if path is None:
            continue
        df = extract_region(path, region, metric)
        base = df if base is None else base.merge(
            df.drop(columns=["sector_name"]), on="sector_code", how="outer")

    if base is None:
        raise ValueError("metric_files must include at least one metric")

    for metric in ("production", "value_added", "employment"):
        for scope in ("in_region", "out_region"):
            col = f"multiplier_{metric}_{scope}"
            if col not in base.columns:
                base[col] = pd.NA
    base["unit_production"] = "원/원"
    base["unit_value_added"] = "원/원"
    base["unit_employment"] = "명/10억원"
    base["region"] = region
    base["table_year"] = table_year
    base["classification"] = classification
    return base
